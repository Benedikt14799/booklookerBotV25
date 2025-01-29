import openpyxl
from openpyxl import Workbook


class ExcelExporter:
    """
    Diese Klasse erstellt bzw. überschreibt eine Excel-Datei (XLSX) und schreibt die
    Überschriften (headers) als Spaltenüberschriften (also in Zeile 1, Spalte A..Z).
    Jeder Datensatz (aus data_rows) wird dann in einer eigenen Zeile darunter ausgegeben.

    Funktionsweise:
    1. self.headers: Liste der gewünschten Spaltenüberschriften.
       - Diese werden in Zeile 1 eingetragen (Spalten 1..n).
    2. self.mapping: Dictionary, das DB-Feldnamen (z.B. "db_autor") auf die entsprechenden
       Spaltenüberschriften (z.B. "C:Autor") abbildet.
    3. export_to_excel(filename, data_rows):
       - Erzeugt eine neue oder überschreibt eine bestehende Excel-Datei.
       - Schreibt self.headers in Zeile 1 (Spalten 1..n).
       - Jeder Datensatz (Dictionary in data_rows) wird in einer eigenen Zeile eingetragen,
         wobei das Mapping entscheidet, in welche Spalte der jeweilige Wert gehört.
    """

    def __init__(self):
        """
        Initialisiert die Excel-Überschriften (self.headers) und die Mapping-Tabelle (self.mapping).
        Die Mapping-Einträge sind als Beispiel hinterlegt, bitte anpassen/anreichern.
        """
        self.headers = [
            "Action(SiteID=Germany|Country=DE|Currency=EUR|Version=1193)",
            "Custom label (SKU)",
            "Category name",
            "Title",
            "Relationship",
            "Relationship details",
            "Schedule Time",
            "P:ISBN",
            "P:EPID",
            "Start price",
            "Quantity",
            "Item photo URL",
            "VideoID",
            "Condition ID",
            "Description",
            "Format",
            "Duration",
            "Buy It Now price",
            "Best Offer Enabled",
            "Best Offer Auto Accept Price",
            "Minimum Best Offer Price",
            "VAT%",
            "Immediate pay required",
            "Location",
            "Shipping service 1 option",
            "Shipping service 1 cost",
            "Shipping service 1 priority",
            "Shipping service 2 option",
            "Shipping service 2 cost",
            "Shipping service 2 priority",
            "Max dispatch time",
            "Returns accepted option",
            "Returns within option",
            "Refund option",
            "Return shipping cost paid by",
            "Shipping profile name",
            "Return profile name",
            "Payment profile name",
            "ProductCompliancePolicyID",
            "Regional ProductCompliancePolicies",
            "C:Autor",
            "C:Buchtitel",
            "C:Sprache",
            "C:Thematik",
            "C:Buchreihe",
            "C:Genre",
            "C:Verlag",
            "C:Erscheinungsjahr",
            "C:Originalsprache",
            "C:Format",
            "C:Herstellungsland und -region",
            "C:Produktart",
            "C:Literarische Gattung",
            "C:Zielgruppe",
            "C:Signiert von",
            "C:Ausgabe",
            "C:Literarische Bewegung",
            "Product Safety Pictograms",
            "Product Safety Statements",
            "Product Safety Component",
            "Regulatory Document Ids",
            "Manufacturer Name",
            "Manufacturer AddressLine1",
            "Manufacturer AddressLine2",
            "Manufacturer City",
            "Manufacturer Country",
            "Manufacturer PostalCode",
            "Manufacturer StateOrProvince",
            "Manufacturer Phone",
            "Manufacturer Email",
            "Manufacturer ContactURL",
            "Responsible Person 1",
            "Responsible Person 1 Type",
            "Responsible Person 1 AddressLine1",
            "Responsible Person 1 AddressLine2",
            "Responsible Person 1 City",
            "Responsible Person 1 Country",
            "Responsible Person 1 PostalCode",
            "Responsible Person 1 StateOrProvince",
            "Responsible Person 1 Phone",
            "Responsible Person 1 Email",
            "Responsible Person 1 ContactURL"
        ]

        # Mapping-Tabelle: DB-Feldname -> Excel-Header
        # Trage hier deine eigenen DB-Feldnamen ein und ent-kommentiere sie.
        # Beispiel:
        # "db_action": "Action(SiteID=Germany|Country=DE|Currency=EUR|Version=1193)",
        # "db_sku": "Custom label (SKU)",
        # "db_category": "Category name",
        # "db_title": "Title",
        # "db_relationship": "Relationship",
        # ...
        self.mapping = {
             "action": "Action(SiteID=Germany|Country=DE|Currency=EUR|Version=1193)",
             "id": "Custom label (SKU)",
             "categoryname": "Category name",
             "title": "Title",
            # "db_relationship": "Relationship",
            # "db_relationship_details": "Relationship details",
            # "db_schedule_time": "Schedule Time",
             "isbn": "P:ISBN",
            # "db_epid": "P:EPID",
             "start_price": "Start price",
            # "db_quantity": "Quantity",
            # "db_item_photo_url": "Item photo URL",
            # "db_video_id": "VideoID",
            # "db_condition_id": "Condition ID",
            # "db_description": "Description",
            # "db_format": "Format",
            # "db_duration": "Duration",
            # "db_buy_it_now_price": "Buy It Now price",
            # "db_best_offer_enabled": "Best Offer Enabled",
            # "db_best_offer_auto_accept_price": "Best Offer Auto Accept Price",
            # "db_minimum_best_offer_price": "Minimum Best Offer Price",
            # "db_vat_percent": "VAT%",
            # "db_immediate_pay_required": "Immediate pay required",
            # "db_location": "Location",
            # "db_shipping_service_1_option": "Shipping service 1 option",
            # "db_shipping_service_1_cost": "Shipping service 1 cost",
            # "db_shipping_service_1_priority": "Shipping service 1 priority",
            # "db_shipping_service_2_option": "Shipping service 2 option",
            # "db_shipping_service_2_cost": "Shipping service 2 cost",
            # "db_shipping_service_2_priority": "Shipping service 2 priority",
            # "db_max_dispatch_time": "Max dispatch time",
            # "db_returns_accepted_option": "Returns accepted option",
            # "db_returns_within_option": "Returns within option",
            # "db_refund_option": "Refund option",
            # "db_return_shipping_cost_paid_by": "Return shipping cost paid by",
            # "db_shipping_profile_name": "Shipping profile name",
            # "db_return_profile_name": "Return profile name",
            # "db_payment_profile_name": "Payment profile name",
            # "db_product_compliance_policy_id": "ProductCompliancePolicyID",
            # "db_regional_product_compliance_policies": "Regional ProductCompliancePolicies",
            # "db_autor": "C:Autor",
            # "db_buchtitel": "C:Buchtitel",
            # "db_sprache": "C:Sprache",
            # "db_thematik": "C:Thematik",
            # "db_buchreihe": "C:Buchreihe",
            # "db_genre": "C:Genre",
            # "db_verlag": "C:Verlag",
            # "db_erscheinungsjahr": "C:Erscheinungsjahr",
            # "db_originalsprache": "C:Originalsprache",
            # "db_buch_format": "C:Format",
            # "db_herstellungsland_und_region": "C:Herstellungsland und -region",
            # "db_produktart": "C:Produktart",
            # "db_literarische_gattung": "C:Literarische Gattung",
            # "db_zielgruppe": "C:Zielgruppe",
            # "db_signiert_von": "C:Signiert von",
            # "db_ausgabe": "C:Ausgabe",
            # "db_literarische_bewegung": "C:Literarische Bewegung",
            # "db_product_safety_pictograms": "Product Safety Pictograms",
            # "db_product_safety_statements": "Product Safety Statements",
            # "db_product_safety_component": "Product Safety Component",
            # "db_regulatory_document_ids": "Regulatory Document Ids",
            # "db_manufacturer_name": "Manufacturer Name",
            # "db_manufacturer_addressline1": "Manufacturer AddressLine1",
            # "db_manufacturer_addressline2": "Manufacturer AddressLine2",
            # "db_manufacturer_city": "Manufacturer City",
            # "db_manufacturer_country": "Manufacturer Country",
            # "db_manufacturer_postalcode": "Manufacturer PostalCode",
            # "db_manufacturer_stateorprovince": "Manufacturer StateOrProvince",
            # "db_manufacturer_phone": "Manufacturer Phone",
            # "db_manufacturer_email": "Manufacturer Email",
            # "db_manufacturer_contacturl": "Manufacturer ContactURL",
            # "db_responsible_person_1": "Responsible Person 1",
            # "db_responsible_person_1_type": "Responsible Person 1 Type",
            # "db_responsible_person_1_addressline1": "Responsible Person 1 AddressLine1",
            # "db_responsible_person_1_addressline2": "Responsible Person 1 AddressLine2",
            # "db_responsible_person_1_city": "Responsible Person 1 City",
            # "db_responsible_person_1_country": "Responsible Person 1 Country",
            # "db_responsible_person_1_postalcode": "Responsible Person 1 PostalCode",
            # "db_responsible_person_1_stateorprovince": "Responsible Person 1 StateOrProvince",
            # "db_responsible_person_1_phone": "Responsible Person 1 Phone",
            # "db_responsible_person_1_email": "Responsible Person 1 Email",
            # "db_responsible_person_1_contacturl": "Responsible Person 1 ContactURL",
        }


    def export_to_excel(self, filename: str, data_rows: list[dict]) -> None:
        """
        Erstellt eine neue Excel-Datei (oder überschreibt eine bestehende),
        schreibt die Headers in Zeile 1 und füllt ab Zeile 2 die Werte aus data_rows.
        Die Spaltenzuordnung erfolgt über self.mapping.

        :param filename: Name der Excel-Datei (z.B. 'export.xlsx').
        :param data_rows: Liste von Dictionaries mit deinen Daten. Die Keys
                          sind die DB-Feldnamen, die in self.mapping abgebildet
                          werden.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Daten"

        # 1) Schreibe die Headers in Zeile 1 (columns)
        for col_index, header in enumerate(self.headers, start=1):
            ws.cell(row=1, column=col_index, value=header)

        # 2) Baue ein Nachschlage-Dict Header -> Spalte
        header_to_col = {}
        for col_index, header in enumerate(self.headers, start=1):
            header_to_col[header] = col_index

        # 3) Jede dict in data_rows wird in einer neuen Zeile (start=2) eingetragen
        for row_index, record in enumerate(data_rows, start=2):
            for db_field, value in record.items():
                if db_field in self.mapping:
                    excel_header = self.mapping[db_field]
                    col_index = header_to_col.get(excel_header)
                    if col_index is not None:
                        ws.cell(row=row_index, column=col_index, value=value)

        # 4) Speichern
        wb.save(filename)